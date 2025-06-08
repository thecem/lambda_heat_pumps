import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_modbus_address(module_index, subindex, number):
    """Generate Modbus address based on module index and number."""
    return f"{module_index}{number:03d}"

def create_sensor_excel():
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lambda Heat Pump Sensors"

    # Define headers
    headers = [
        "Module Type",
        "Sensor ID",
        "Name",
        "Description",
        "Unit",
        "Data Type",
        "Read/Write",
        "Min Value",
        "Max Value",
        "Scale",
        "Precision",
        "Modbus Address"
    ]

    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Module definitions with Modbus addresses
    modules = {
        "General Ambient": {
            "prefix": "amb",
            "index": 0,
            "subindex": 0,
            "sensors": [
                {"id": "error_number", "name": "Error Number", "description": "Error code", "unit": None, "data_type": "int16", "rw": "RO", "scale": 1, "precision": 0, "number": 0},
                {"id": "operating_state", "name": "Operating State", "description": "Current operating state", "unit": None, "data_type": "uint16", "rw": "RO", "scale": 1, "precision": 0, "number": 1},
                {"id": "actual_ambient_temp", "name": "Actual Ambient Temperature", "description": "Current ambient temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": -50.0, "max": 80.0, "number": 2},
                {"id": "average_ambient_temp_1h", "name": "Average Ambient Temperature 1h", "description": "Average temperature of the last 60 minutes", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 3},
                {"id": "calculated_ambient_temp", "name": "Calculated Ambient Temperature", "description": "Calculated temperature for heat distribution modules", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 4}
            ]
        },
        "Heat Pump": {
            "prefix": "hp",
            "index": 1,
            "subindex": 0,
            "sensors": [
                {"id": "error_state", "name": "Error State", "description": "Current error state", "unit": None, "data_type": "uint16", "rw": "RO", "scale": 1, "precision": 0, "number": 0},
                {"id": "error_number", "name": "Error Number", "description": "Error code", "unit": None, "data_type": "int16", "rw": "RO", "scale": 1, "precision": 0, "number": 1},
                {"id": "state", "name": "State", "description": "Current state", "unit": None, "data_type": "uint16", "rw": "RO", "scale": 1, "precision": 0, "number": 2},
                {"id": "operating_state", "name": "Operating State", "description": "Current operating state", "unit": None, "data_type": "uint16", "rw": "RO", "scale": 1, "precision": 0, "number": 3},
                {"id": "flow_temp", "name": "Flow Temperature", "description": "Flow line temperature", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.01, "precision": 2, "number": 4},
                {"id": "return_temp", "name": "Return Temperature", "description": "Return line temperature", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.01, "precision": 2, "number": 5},
                {"id": "volume_flow_sink", "name": "Volume Flow Sink", "description": "Volume flow heat sink", "unit": "l/min", "data_type": "int16", "rw": "RO", "scale": 0.01, "precision": 2, "number": 6},
                {"id": "source_inlet_temp", "name": "Source Inlet Temperature", "description": "Energy source inlet temperature", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.01, "precision": 2, "number": 7},
                {"id": "source_outlet_temp", "name": "Source Outlet Temperature", "description": "Energy source outlet temperature", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.01, "precision": 2, "number": 8},
                {"id": "volume_flow_source", "name": "Volume Flow Source", "description": "Volume flow energy source", "unit": "l/min", "data_type": "int16", "rw": "RO", "scale": 0.01, "precision": 2, "number": 9},
                {"id": "compressor_rating", "name": "Compressor Rating", "description": "Compressor power rating", "unit": "%", "data_type": "uint16", "rw": "RO", "scale": 0.01, "precision": 2, "number": 10},
                {"id": "heating_power", "name": "Heating Power", "description": "Current heating power", "unit": "kW", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 11},
                {"id": "power_consumption", "name": "Power Consumption", "description": "Frequency inverter power consumption", "unit": "W", "data_type": "int16", "rw": "RO", "scale": 1, "precision": 0, "number": 12},
                {"id": "cop", "name": "COP", "description": "Coefficient of Performance", "unit": None, "data_type": "int16", "rw": "RO", "scale": 0.01, "precision": 2, "number": 13},
                {"id": "energy_consumption", "name": "Energy Consumption", "description": "Electric energy consumption since last reset", "unit": "Wh", "data_type": "int32", "rw": "RO", "scale": 1, "precision": 0, "number": 20},
                {"id": "energy_output", "name": "Energy Output", "description": "Thermal energy output since last reset", "unit": "Wh", "data_type": "int32", "rw": "RO", "scale": 1, "precision": 0, "number": 22}
            ]
        },
        "Boiler": {
            "prefix": "boil",
            "index": 2,
            "subindex": 0,
            "sensors": [
                {"id": "error_number", "name": "Error Number", "description": "Error code", "unit": None, "data_type": "int16", "rw": "RO", "scale": 1, "precision": 0, "number": 0},
                {"id": "operating_state", "name": "Operating State", "description": "Current operating state", "unit": None, "data_type": "uint16", "rw": "RO", "scale": 1, "precision": 0, "number": 1},
                {"id": "actual_high_temp", "name": "Actual High Temperature", "description": "Current temperature upper sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 2},
                {"id": "actual_low_temp", "name": "Actual Low Temperature", "description": "Current temperature lower sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 3},
                {"id": "circulation_temp", "name": "Circulation Temperature", "description": "Current temperature circulation sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 4},
                {"id": "circulation_pump_state", "name": "Circulation Pump State", "description": "Current state of circulation pump", "unit": None, "data_type": "int16", "rw": "RO", "scale": 1, "precision": 0, "number": 5},
                {"id": "max_boiler_temp", "name": "Maximum Boiler Temperature", "description": "Setting for maximum boiler temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 25.0, "max": 65.0, "number": 50}
            ]
        },
        "Buffer": {
            "prefix": "buff",
            "index": 3,
            "subindex": 0,
            "sensors": [
                {"id": "error_number", "name": "Error Number", "description": "Error code", "unit": None, "data_type": "int16", "rw": "RO", "scale": 1, "precision": 0, "number": 0},
                {"id": "operating_state", "name": "Operating State", "description": "Current operating state", "unit": None, "data_type": "uint16", "rw": "RO", "scale": 1, "precision": 0, "number": 1},
                {"id": "actual_high_temp", "name": "Actual High Temperature", "description": "Current temperature upper sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 2},
                {"id": "actual_low_temp", "name": "Actual Low Temperature", "description": "Current temperature lower sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 3},
                {"id": "modbus_buffer_temp_high", "name": "High Temperature (Modbus)", "description": "Buffer temperature set via Modbus", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 0.0, "max": 90.0, "number": 4},
                {"id": "request_type", "name": "Request Type", "description": "Type of current request", "unit": None, "data_type": "int16", "rw": "RW", "scale": 1, "precision": 0, "number": 5},
                {"id": "request_flow_line_temp", "name": "Request Flow Line Temperature", "description": "Requested flow line temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 0.0, "max": 65.0, "number": 6},
                {"id": "request_return_line_temp", "name": "Request Return Line Temperature", "description": "Requested return line temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 0.0, "max": 60.0, "number": 7},
                {"id": "request_heat_sink_temp_diff", "name": "Request Heat Sink Temperature Difference", "description": "Requested temperature difference", "unit": "K", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 0.0, "max": 35.0, "number": 8},
                {"id": "request_heating_capacity", "name": "Request Heating Capacity", "description": "Requested power", "unit": "kW", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 0.0, "max": 25.5, "number": 9},
                {"id": "max_buffer_temp", "name": "Maximum Buffer Temperature", "description": "Setting for maximum buffer temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 25.0, "max": 65.0, "number": 50}
            ]
        },
        "Solar": {
            "prefix": "sol",
            "index": 4,
            "subindex": 0,
            "sensors": [
                {"id": "error_number", "name": "Error Number", "description": "Error code", "unit": None, "data_type": "int16", "rw": "RO", "scale": 1, "precision": 0, "number": 0},
                {"id": "operating_state", "name": "Operating State", "description": "Current operating state", "unit": None, "data_type": "uint16", "rw": "RO", "scale": 1, "precision": 0, "number": 1},
                {"id": "collector_temp", "name": "Collector Temperature", "description": "Current temperature collector sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 2},
                {"id": "buffer1_temp", "name": "Buffer 1 Temperature", "description": "Current temperature buffer 1 sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 3},
                {"id": "buffer2_temp", "name": "Buffer 2 Temperature", "description": "Current temperature buffer 2 sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 4},
                {"id": "max_buffer_temp", "name": "Maximum Buffer Temperature", "description": "Setting for maximum buffer temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 25.0, "max": 90.0, "number": 50},
                {"id": "buffer_changeover_temp", "name": "Buffer Changeover Temperature", "description": "Setting for buffer changeover temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 25.0, "max": 90.0, "number": 51}
            ]
        },
        "Heating Circuit": {
            "prefix": "hc",
            "index": 5,
            "subindex": 0,
            "sensors": [
                {"id": "error_number", "name": "Error Number", "description": "Error code", "unit": None, "data_type": "int16", "rw": "RO", "scale": 1, "precision": 0, "number": 0},
                {"id": "operating_state", "name": "Operating State", "description": "Current operating state", "unit": None, "data_type": "uint16", "rw": "RO", "scale": 1, "precision": 0, "number": 1},
                {"id": "flow_line_temp", "name": "Flow Line Temperature", "description": "Current temperature flow line sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 2},
                {"id": "return_line_temp", "name": "Return Line Temperature", "description": "Current temperature return line sensor", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 3},
                {"id": "room_device_temp", "name": "Room Device Temperature", "description": "Current temperature room device sensor", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": -29.9, "max": 99.9, "number": 4},
                {"id": "set_flow_line_temp", "name": "Set Flow Line Temperature", "description": "Target flow line temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 15.0, "max": 65.0, "number": 5},
                {"id": "operating_mode", "name": "Operating Mode", "description": "Current mode of operation", "unit": None, "data_type": "int16", "rw": "RW", "scale": 1, "precision": 0, "number": 6},
                {"id": "target_flow_line_temp", "name": "Target Flow Line Temperature", "description": "Calculated target flow temperature", "unit": "°C", "data_type": "int16", "rw": "RO", "scale": 0.1, "precision": 1, "number": 7},
                {"id": "set_flow_line_offset_temp", "name": "Set Flow Line Offset Temperature", "description": "Offset for flow temperature", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": -10.0, "max": 10.0, "number": 50},
                {"id": "set_room_heating_temp", "name": "Set Room Heating Temperature", "description": "Setting for room temperature setpoint heating operation", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 15.0, "max": 40.0, "number": 51},
                {"id": "set_room_cooling_temp", "name": "Set Room Cooling Temperature", "description": "Setting for room temperature setpoint cooling operation", "unit": "°C", "data_type": "int16", "rw": "RW", "scale": 0.1, "precision": 1, "min": 15.0, "max": 40.0, "number": 52}
            ]
        }
    }

    # Write data
    row = 2
    for module_name, module_data in modules.items():
        for sensor in module_data["sensors"]:
            ws.cell(row=row, column=1).value = module_name
            ws.cell(row=row, column=2).value = f"{module_data['prefix']}_{sensor['id']}"
            ws.cell(row=row, column=3).value = sensor["name"]
            ws.cell(row=row, column=4).value = sensor["description"]
            ws.cell(row=row, column=5).value = sensor["unit"]
            ws.cell(row=row, column=6).value = sensor["data_type"]
            ws.cell(row=row, column=7).value = sensor["rw"]
            ws.cell(row=row, column=8).value = sensor.get("min")
            ws.cell(row=row, column=9).value = sensor.get("max")
            ws.cell(row=row, column=10).value = sensor["scale"]
            ws.cell(row=row, column=11).value = sensor["precision"]
            # Generate and write Modbus address
            modbus_address = generate_modbus_address(module_data["index"], module_data["subindex"], sensor["number"])
            ws.cell(row=row, column=12).value = modbus_address
            row += 1

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save the workbook
    wb.save("lambda_heat_pump_sensors.xlsx")

if __name__ == "__main__":
    create_sensor_excel() 